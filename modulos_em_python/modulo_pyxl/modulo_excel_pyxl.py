"""
https://openpyxl.readthedocs.io/en/stable/

"""

import openpyxl
from random import uniform

analise = openpyxl.load_workbook('analise_sol.xlsx')
nome_planilhas = analise.sheetnames  # ver as abas
# print(nome_planilhas)
planilha1 = analise['Minimum template']  # armazena a aba 'minimum template
'''
print(planilha1['b4'])  # objeto
print(planilha1['b4'].value)  # valor
print(planilha1['b'])  # tupla com objetos da coluna = posso iterar

for campo in planilha1['b']:  # uma coluna da aba
    print(campo.value)

for linha in planilha1['a1:c2']:  # uma 'zona' da tabela
    for coluna in linha:
        print(coluna.value)

for linha in planilha1:  # toda a planilha
    for coluna in linha:
        print(coluna.value)
'''
'''
for linha in planilha1:  # mostrar em colunas no temrinal
    print(linha[0].value, linha[1].value, linha[2].value, linha[3].value, linha[4].value)

# posso editar a planilha, mas preciso salvvar a nova planilha, nao vai modif a orgnl
planilha1['c5'].value = 38

for linha in range(13, 24):
    planilha1.cell(linha, 1).value = 'TESTE'  # escrevendo na celula com a coordenada cell(x,y)

analise.save('nova_analise.xlsx')
'''

# criando uma planilha
planilha = openpyxl.Workbook()
# criar aba dentro da planilha
planilha.create_sheet('Planilha1', 0)
planilha.create_sheet('Planilha2', 1)

planilhaA = planilha['Planilha1']
planilhaB = planilha['Planilha2']

for linha in range(1, 11):
    numero_pedido = linha - 1
    planilhaA.cell(linha, 1).value = numero_pedido
    planilhaA.cell(linha, 2).value = 1200 + linha

    preco = round(uniform(10, 100), 2)
    planilhaA.cell(linha, 3).value = preco

for linha in range(1, 11):
    planilhaB.cell(linha, 1).value = f'martin {linha} {round(uniform(10,100),2)}'
    planilhaB.cell(linha, 2).value = f'lina {linha} {round(uniform(10,100),2)}'
    planilhaB.cell(linha, 3).value = f'gabriela {linha} {round(uniform(10,100),2)}'


planilha.save('planilha_criada.xlsx')
